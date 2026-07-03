"""Excel (.xlsx) document processor using openpyxl."""

import asyncio
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

MAX_ROWS = 100_000


def _load_workbook_fn():
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        raise ImportError("openpyxl is required for XLSX processing. Install with: pip install openpyxl") from None
    return load_workbook


class XLSXProcessor:
    """Extract text and metadata from Excel files."""

    async def extract_stub(self, file_path: str) -> dict:
        """Discovery stub for a tabular file — sheet names, column headers, and
        a row count (not the full table). Best-effort: minimal stub on failure."""
        load_workbook = _load_workbook_fn()

        def _extract() -> dict:
            try:
                wb = load_workbook(file_path, data_only=True, read_only=True)
            except Exception as e:  # noqa: BLE001
                logger.debug("XLSX stub extraction failed for %s: %s", file_path, e)
                return {"sheets": [], "summary": "Excel file (summary unavailable)"}
            try:
                sheets = []
                for name in wb.sheetnames:
                    ws = wb[name]
                    if ws.sheet_state != "visible":
                        continue
                    headers = [str(c.value).strip() for c in next(ws.iter_rows(max_row=1), ()) if c.value is not None]
                    sheets.append({"name": name, "columns": headers, "row_count": ws.max_row})
                names = ", ".join(s["name"] for s in sheets)
                cols = ", ".join(c for s in sheets for c in s["columns"])
                summary = (
                    f"Excel workbook — sheet(s): {names}. Columns: {cols}." if sheets else "Excel workbook (empty)."
                )
                return {"sheets": sheets, "summary": summary}
            finally:
                wb.close()

        return await asyncio.to_thread(_extract)

    async def extract_text(self, file_path: str) -> str:
        """Extract text from an XLSX file."""
        load_workbook = _load_workbook_fn()

        def _extract():
            try:
                wb = load_workbook(file_path, data_only=True, read_only=True)
            except Exception as e:
                if "password" in str(e).lower() or "encrypted" in str(e).lower():
                    raise ValueError("Password-protected Excel files are not supported") from e
                raise

            text_parts = []

            for sheet in wb.sheetnames:
                ws = wb[sheet]
                if ws.sheet_state != "visible":
                    continue

                text_parts.append(f"--- Sheet: {sheet} ---")

                for row_count, row in enumerate(ws.iter_rows(), start=1):
                    if row_count > MAX_ROWS:
                        logger.warning("Sheet '%s' exceeds %d rows, truncating", sheet, MAX_ROWS)
                        text_parts.append(f"[Truncated: exceeded {MAX_ROWS} rows]")
                        break

                    cells = [str(cell.value) if cell.value is not None else "" for cell in row]
                    row_text = "\t".join(cells)
                    if row_text.strip():
                        text_parts.append(row_text)

            wb.close()
            return "\n".join(text_parts)

        return await asyncio.to_thread(_extract)

    async def extract_metadata(self, file_path: str) -> dict:
        """Best-effort metadata extraction — returns {} on any failure."""
        load_workbook = _load_workbook_fn()

        def _extract():
            try:
                wb = load_workbook(file_path, read_only=True)
                try:
                    props = getattr(wb, "properties", None)
                    if props is None:
                        return {}
                    author = (getattr(props, "creator", None) or "").strip() or None
                    title = (getattr(props, "title", None) or "").strip() or None
                    created = getattr(props, "created", None)
                    authored_at = created if isinstance(created, datetime) else None
                    return {"author": author, "doc_title": title, "authored_at": authored_at}
                finally:
                    wb.close()
            except Exception as e:
                logger.debug("XLSX metadata extraction failed for %s: %s", file_path, e)
                return {}

        return await asyncio.to_thread(_extract)
